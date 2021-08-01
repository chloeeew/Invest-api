"""
==================
Author:Chloeee
Time:2021/5/11 21:54
Contact:403505960@qq.com
==================
"""

from config.path import cases_file
from common.logger_handler import logger
import openpyxl

"""封装处理打开excel读取数据的操作"""


class ExcelManager:
    def __init__(self,sheet_name):
        self.wb = openpyxl.load_workbook(cases_file)
        self.m_sheet = self.wb[sheet_name]
        logger.info(f"\n\n\n打开表单{sheet_name}\n\n")

    def get_excel_data(self) -> list:
        """
        获取sheet中的所有数据，第一行作为key，后续行作为values，每一行数据是字典形式，用列表存储
        返回列表
        """
        # 最大行数
        max_row = self.m_sheet.max_row
        # 最大列数
        max_column = self.m_sheet.max_column
        keys = []   # 存储第一行作为key
        for t in range(1, max_column + 1):
            keys.append(self.m_sheet.cell(1, t).value)

        values = []     # 存储第二行及之后的数据
        results = []   # 存储所有结果
        for r in range(2, max_row + 1):
            for c in range(1, max_column + 1):
                m_value = self.m_sheet.cell(r, c).value
                values.append(m_value)
            result_dict = dict(zip(keys, values))
            results.append(result_dict)
            values = []
        return results



if __name__ == "__main__":
    print(ExcelManager("充值接口").get_excel_data())
